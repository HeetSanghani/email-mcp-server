import sys
import os
import json
import re
from pathlib import Path
from typing import Optional
from datetime import datetime
from typing import Optional

# Add project root to path so we can import auth module
sys.path.append(str(Path(__file__).parent.parent.absolute()))

from mcp.server.fastmcp import FastMCP
from auth.gmail_auth import (
    is_authenticated as gmail_is_auth,
    gmail_list_emails,
    gmail_search_emails,
    gmail_get_email,
    gmail_send_email,
    gmail_create_draft,
    gmail_reply_to_email,
    gmail_mark_as_read,
    gmail_mark_as_unread,
    gmail_archive_email,
    gmail_delete_email,
    gmail_forward_email,
)
from auth.outlook_auth import (
    is_authenticated as outlook_is_auth,
    outlook_list_emails,
    outlook_search_emails,
    outlook_get_email,
    outlook_send_email,
    outlook_create_draft,
    outlook_reply_to_email,
    outlook_mark_as_read,
    outlook_mark_as_unread,
    outlook_archive_email,
    outlook_delete_email,
    outlook_forward_email,
)

mcp = FastMCP("Email-MCP-Server")


# ─────────────────────────────────────────────
# EXISTING TOOLS (unchanged)
# ─────────────────────────────────────────────

@mcp.tool()
def list_recent_emails(provider: str = None, max_results: int = 10) -> list:
    """
    List recent emails from Gmail, Outlook, or both.

    Parameters:
    - provider: "gmail", "outlook", or None (to fetch from all connected providers)
    - max_results: Maximum number of emails to fetch (default: 10)
    """
    emails = []
    provider = provider.lower().strip() if provider else None

    if provider == "gmail" or not provider:
        if gmail_is_auth():
            try:
                emails.extend(gmail_list_emails(max_results))
            except Exception as e:
                print(f"Error fetching from Gmail: {e}", file=sys.stderr)
        elif provider == "gmail":
            raise RuntimeError("Gmail is not authenticated.")

    if provider == "outlook" or not provider:
        if outlook_is_auth():
            try:
                emails.extend(outlook_list_emails(max_results))
            except Exception as e:
                print(f"Error fetching from Outlook: {e}", file=sys.stderr)
        elif provider == "outlook":
            raise RuntimeError("Outlook is not authenticated.")

    return emails


@mcp.tool()
def search_emails(query: str, provider: str = None, max_results: int = 10) -> list:
    """
    Search emails in Gmail, Outlook, or both.

    Parameters:
    - query: Search query string (e.g. sender name, keyword, or phrase)
    - provider: "gmail", "outlook", or None (to search all connected providers)
    - max_results: Maximum number of search results (default: 10)
    """
    emails = []
    provider = provider.lower().strip() if provider else None

    if provider == "gmail" or not provider:
        if gmail_is_auth():
            try:
                emails.extend(gmail_search_emails(query, max_results))
            except Exception as e:
                print(f"Error searching Gmail: {e}", file=sys.stderr)
        elif provider == "gmail":
            raise RuntimeError("Gmail is not authenticated.")

    if provider == "outlook" or not provider:
        if outlook_is_auth():
            try:
                emails.extend(outlook_search_emails(query, max_results))
            except Exception as e:
                print(f"Error searching Outlook: {e}", file=sys.stderr)
        elif provider == "outlook":
            raise RuntimeError("Outlook is not authenticated.")

    return emails


@mcp.tool()
def get_email_details(email_id: str, provider: str) -> dict:
    """
    Get the full details of a specific email including body content.

    Parameters:
    - email_id: The unique ID of the email
    - provider: The provider of the email ("gmail" or "outlook")
    """
    provider = provider.lower().strip()
    if provider == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        return gmail_get_email(email_id)
    elif provider == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        return outlook_get_email(email_id)
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def send_email(to: str, subject: str, body: str, provider: str, attachment_paths: Optional[list[str]] = None) -> str:
    """
    Send a new email using Gmail or Outlook, with optional attachments.

    Parameters:
    - to: Recipient's email address
    - subject: Subject of the email
    - body: Plain text body of the email
    - provider: The provider to send the email with ("gmail" or "outlook")
    - attachment_paths: Optional list of absolute file paths to attach
    """
    provider = provider.lower().strip()
    if provider == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_send_email(to, subject, body, attachment_paths=attachment_paths)
        return "Email sent successfully via Gmail."
    elif provider == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_send_email(to, subject, body, attachment_paths=attachment_paths)
        return "Email sent successfully via Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def create_draft(to: str, subject: str, body: str, provider: str) -> str:
    """
    Create a new draft email in Gmail or Outlook.

    Parameters:
    - to: Recipient's email address
    - subject: Subject of the email
    - body: Plain text body of the email
    - provider: The provider to create the draft in ("gmail" or "outlook")
    """
    provider = provider.lower().strip()
    if provider == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_create_draft(to, subject, body)
        return "Draft created successfully in Gmail."
    elif provider == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_create_draft(to, subject, body)
        return "Draft created successfully in Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def reply_to_email(email_id: str, body: str, provider: str) -> str:
    """
    Reply to an existing email thread.

    Parameters:
    - email_id: The ID of the email to reply to
    - body: The body of your reply message
    - provider: The email provider ("gmail" or "outlook")
    """
    provider = provider.lower().strip()
    if provider == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_reply_to_email(email_id, body)
        return "Reply sent successfully via Gmail."
    elif provider == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_reply_to_email(email_id, body)
        return "Reply sent successfully via Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


# ─────────────────────────────────────────────
# NEW TOOL 1 — categorize_inbox
# ─────────────────────────────────────────────

@mcp.tool()
def categorize_inbox(provider: str, max_results: int = 15) -> dict:
    """
    Fetch emails and automatically categorize them using AI into:
    Important, Spam, Purchases/Orders, Newsletters, Social,
    Work, Security Alerts, and Promotions.

    Parameters:
    - provider: "gmail", "outlook", or "all"
    - max_results: How many emails to categorize (default: 50)

    Returns a dict with category names as keys and lists of emails as values.
    """
    # ── 1. Fetch emails ──────────────────────────────────────────────
    # Hard-cap at 20 — larger batches exceed Groq's TPM token limit
    max_results = min(max_results, 20)
    emails = []
    provider = provider.lower().strip()

    if provider in ("gmail", "all"):
        if gmail_is_auth():
            try:
                emails.extend(gmail_list_emails(max_results))
            except Exception as e:
                print(f"Gmail fetch error: {e}", file=sys.stderr)
        elif provider == "gmail":
            raise RuntimeError("Gmail is not authenticated.")

    if provider in ("outlook", "all"):
        if outlook_is_auth():
            try:
                emails.extend(outlook_list_emails(max_results))
            except Exception as e:
                print(f"Outlook fetch error: {e}", file=sys.stderr)
        elif provider == "outlook":
            raise RuntimeError("Outlook is not authenticated.")

    if not emails:
        return {"error": "No emails fetched. Check authentication."}

    # ── 2. Build prompt for Claude ───────────────────────────────────
    categories = [
        "Important",
        "Spam",
        "Purchases/Orders",
        "Newsletters",
        "Social",
        "Work",
        "Security Alerts",
        "Promotions",
    ]

    email_summaries = []
    for i, email in enumerate(emails):
        email_summaries.append(
            f"[{i}] From: {email.get('from', 'N/A')[:40]} | "
            f"Subject: {email.get('subject', 'N/A')[:60]} | "
            f"Preview: {email.get('snippet', '')[:50]}"
        )

    prompt = (
        f"Categorize {len(emails)} emails into ONE of: "
        + ", ".join(categories)
        + "\n\nRules: Security Alerts=password/login alerts; Purchases/Orders=receipts/shipping; "
        "Newsletters=digests/blogs; Promotions=sales/discounts; Social=social media; "
        "Work=colleagues/clients; Spam=unsolicited; Important=everything else.\n\n"
        "Emails:\n" + "\n".join(email_summaries)
        + "\n\nRespond ONLY with valid JSON. Keys=category names, values=arrays of email indexes (integers). "
        f"Include ALL {len(emails)} indexes. No text outside JSON."
    )

    # ── 3. Call AI API (Anthropic → Groq → OpenAI fallback) ─────────
    raw = None

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY") or _load_env_key("ANTHROPIC_API_KEY")
    groq_key      = os.environ.get("GROQ_API_KEY")      or _load_env_key("GROQ_API_KEY")
    openai_key    = os.environ.get("OPENAI_API_KEY")    or _load_env_key("OPENAI_API_KEY")

    placeholder = lambda k: not k or k in ("", "your_anthropic_key_here", "your_groq_key_here", "your_openai_api_key_here")

    if not placeholder(anthropic_key):
        import anthropic as _anthropic
        _client = _anthropic.Anthropic(api_key=anthropic_key)
        _msg = _client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = _msg.content[0].text.strip()
    elif not placeholder(groq_key):
        from openai import OpenAI as _OpenAI
        _client = _OpenAI(api_key=groq_key, base_url="https://api.groq.com/openai/v1")
        _resp = _client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = _resp.choices[0].message.content.strip()
    elif not placeholder(openai_key):
        from openai import OpenAI as _OpenAI
        _client = _OpenAI(api_key=openai_key)
        _resp = _client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = _resp.choices[0].message.content.strip()
    else:
        raise RuntimeError(
            "No AI API key found. Set ANTHROPIC_API_KEY, GROQ_API_KEY, or OPENAI_API_KEY in your .env file."
        )

    raw = raw.strip()

    # Strip markdown fences if present
    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)

    try:
        index_map: dict = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Claude returned invalid JSON: {e}\nRaw: {raw}")

    # ── 4. Build result: category → list of full email dicts ─────────
    result = {cat: [] for cat in categories}
    result["Uncategorized"] = []

    for category, indexes in index_map.items():
        target = category if category in result else "Uncategorized"
        for idx in indexes:
            if 0 <= idx < len(emails):
                result[target].append(emails[idx])

    # Add summary counts at top level
    result["_summary"] = {
        cat: len(result[cat]) for cat in categories + ["Uncategorized"]
    }

    return result


def _load_env_key(key_name: str) -> str:
    """Helper: read a key from .env file in project root."""
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line.startswith(f"{key_name}="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    return ""


# ─────────────────────────────────────────────
# NEW TOOL 2 — export_emails_to_excel
# ─────────────────────────────────────────────

@mcp.tool()
def export_emails_to_excel(
    provider: str,
    max_results: int = 50,
    query: Optional[str] = None,
    save_folder: Optional[str] = None,
) -> dict:
    """
    Export emails to a formatted Excel (.xlsx) file.

    Columns: Email ID, Sender Name, Sender Email, Subject,
             Date Sent/Received, Snippet/Preview, Full Body, Category.

    The file is saved locally AND a base64 version is returned
    so the frontend can offer a browser download.

    Parameters:
    - provider: "gmail", "outlook", or "all"
    - max_results: Number of emails to export (default: 50)
    - query: Optional search query to filter emails (e.g. "from:google.com")
    - save_folder: Local folder to save the file (default: ~/Downloads)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    import base64

    provider_norm = provider.lower().strip()

    # ── 1. Fetch emails ──────────────────────────────────────────────
    emails = []

    if provider_norm in ("gmail", "all"):
        if gmail_is_auth():
            try:
                if query:
                    emails.extend(gmail_search_emails(query, max_results))
                else:
                    emails.extend(gmail_list_emails(max_results))
            except Exception as e:
                print(f"Gmail fetch error: {e}", file=sys.stderr)
        elif provider_norm == "gmail":
            raise RuntimeError("Gmail is not authenticated.")

    if provider_norm in ("outlook", "all"):
        if outlook_is_auth():
            try:
                if query:
                    emails.extend(outlook_search_emails(query, max_results))
                else:
                    emails.extend(outlook_list_emails(max_results))
            except Exception as e:
                print(f"Outlook fetch error: {e}", file=sys.stderr)
        elif provider_norm == "outlook":
            raise RuntimeError("Outlook is not authenticated.")

    if not emails:
        return {"error": "No emails found to export."}

    # ── 2. Parse sender name & email ────────────────────────────────
    def parse_sender(from_str: str):
        """'John Doe <john@example.com>' → ('John Doe', 'john@example.com')"""
        if not from_str:
            return ("", "")
        match = re.match(r"^(.*?)\s*<([^>]+)>$", from_str.strip())
        if match:
            return match.group(1).strip().strip('"'), match.group(2).strip()
        if "@" in from_str:
            return ("", from_str.strip())
        return (from_str.strip(), "")

    # ── 3. Build Excel workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"

    # Header style
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alternate row fill
    alt_fill = PatternFill("solid", start_color="EBF3FB")

    headers = [
        "Email ID",
        "Sender Name",
        "Sender Email",
        "Subject",
        "Date Sent/Received",
        "Snippet / Preview",
        "Full Body",
        "Provider",
    ]

    col_widths = [24, 22, 30, 40, 22, 50, 60, 10]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Write email rows
    for row_idx, email in enumerate(emails, start=2):
        sender_name, sender_email = parse_sender(email.get("from", ""))

        row_data = [
            email.get("id", ""),
            sender_name,
            sender_email,
            email.get("subject", ""),
            email.get("date", ""),
            email.get("snippet", ""),
            email.get("body", email.get("snippet", "")),
            email.get("provider", provider_norm),
        ]

        use_alt = (row_idx % 2 == 0)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx in (6, 7)),  # wrap snippet & body
            )
            cell.border = cell_border
            if use_alt:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 18

    # ── 4. Summary sheet ─────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Export Summary"
    ws_summary["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_summary["A3"] = "Total Emails Exported"
    ws_summary["B3"] = len(emails)
    ws_summary["A3"].font = Font(name="Arial", bold=True)
    ws_summary["A4"] = "Provider"
    ws_summary["B4"] = provider
    ws_summary["A4"].font = Font(name="Arial", bold=True)
    ws_summary["A5"] = "Search Query"
    ws_summary["B5"] = query or "All emails"
    ws_summary["A5"].font = Font(name="Arial", bold=True)
    ws_summary["A6"] = "Exported At"
    ws_summary["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A6"].font = Font(name="Arial", bold=True)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    # ── 5. Save to disk ──────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"emails_export_{timestamp}.xlsx"

    if not save_folder:
        save_folder = str(Path.home() / "Downloads")

    os.makedirs(save_folder, exist_ok=True)
    file_path = os.path.join(save_folder, filename)
    wb.save(file_path)

    # ── 6. Also encode as base64 for frontend download ───────────────
    with open(file_path, "rb") as f:
        file_bytes = f.read()
    b64 = base64.b64encode(file_bytes).decode("utf-8")

    return {
        "success": True,
        "filename": filename,
        "saved_to": file_path,
        "total_emails": len(emails),
        "sheets": ["Emails", "Summary"],
        "base64_xlsx": b64,
        "message": (
            f"Exported {len(emails)} emails to '{file_path}'. "
            "The base64_xlsx field can be used by the frontend to offer a browser download."
        ),
    }


@mcp.tool()
def mark_email_as_read(email_id: str, provider: str) -> str:
    """
    Mark an email as read.

    Parameters:
    - email_id: The ID of the email to mark as read
    - provider: "gmail" or "outlook"
    """
    p = provider.lower().strip()
    if p == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_mark_as_read(email_id)
        return "Email marked as read in Gmail."
    elif p == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_mark_as_read(email_id)
        return "Email marked as read in Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def mark_email_as_unread(email_id: str, provider: str) -> str:
    """
    Mark an email as unread.

    Parameters:
    - email_id: The ID of the email to mark as unread
    - provider: "gmail" or "outlook"
    """
    p = provider.lower().strip()
    if p == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_mark_as_unread(email_id)
        return "Email marked as unread in Gmail."
    elif p == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_mark_as_unread(email_id)
        return "Email marked as unread in Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def archive_email(email_id: str, provider: str) -> str:
    """
    Archive an email (removes it from the inbox).

    Parameters:
    - email_id: The ID of the email to archive
    - provider: "gmail" or "outlook"
    """
    p = provider.lower().strip()
    if p == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_archive_email(email_id)
        return "Email archived (removed from Inbox) in Gmail."
    elif p == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_archive_email(email_id)
        return "Email archived (moved to Archive folder) in Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def delete_email(email_id: str, provider: str) -> str:
    """
    Delete an email (moves it to Trash/Deleted Items).

    Parameters:
    - email_id: The ID of the email to delete
    - provider: "gmail" or "outlook"
    """
    p = provider.lower().strip()
    if p == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_delete_email(email_id)
        return "Email moved to Trash in Gmail."
    elif p == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_delete_email(email_id)
        return "Email moved to Deleted Items in Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


@mcp.tool()
def forward_email(email_id: str, provider: str, to: str, comment: Optional[str] = "") -> str:
    """
    Forward an email to another recipient.

    Parameters:
    - email_id: The ID of the email to forward
    - provider: "gmail" or "outlook"
    - to: Recipient's email address to forward to
    - comment: Optional message/note to prepend to the forwarded email
    """
    p = provider.lower().strip()
    if p == "gmail":
        if not gmail_is_auth():
            raise RuntimeError("Gmail is not authenticated.")
        gmail_forward_email(email_id, to, comment)
        return f"Email forwarded successfully to {to} via Gmail."
    elif p == "outlook":
        if not outlook_is_auth():
            raise RuntimeError("Outlook is not authenticated.")
        outlook_forward_email(email_id, to, comment)
        return f"Email forwarded successfully to {to} via Outlook."
    else:
        raise ValueError("Invalid provider specified. Must be 'gmail' or 'outlook'.")


# ─────────────────────────────────────────────
# MCP RESOURCES (Exposes email data as readable context resources)
# ─────────────────────────────────────────────

@mcp.resource("email://{provider}/{email_id}")
def get_email_body_resource(provider: str, email_id: str) -> str:
    """
    Exposes the raw text content of a specific email as an MCP resource.
    Useful when the AI wants to read the full body content.
    """
    try:
        details = get_email_details(email_id, provider)
        subject = details.get("subject", "No Subject")
        sender = details.get("from", "Unknown")
        body = details.get("body", "")
        return f"From: {sender}\nSubject: {subject}\n\n{body}"
    except Exception as e:
        return f"Error loading email: {str(e)}"


# ─────────────────────────────────────────────
# MCP PROMPTS (Predefined templates for Cline / clients)
# ─────────────────────────────────────────────

@mcp.prompt()
def summarize_inbox(provider: str = "all", count: int = 5) -> str:
    """
    A template prompt to analyze recent emails and highlight action items.
    """
    return (
        f"Please list the {count} most recent emails from provider '{provider}' "
        "and perform the following tasks:\n"
        "1. Summarize the main topic of each email in one sentence.\n"
        "2. Identify if there are any immediate action items or deadlines.\n"
        "3. Draft a brief recommended reply if action is required."
    )


@mcp.prompt()
def write_professional_reply(email_id: str, provider: str, user_instructions: str) -> str:
    """
    A template to draft a clean, professional email reply.
    """
    return (
        f"Fetch the email details for ID '{email_id}' from provider '{provider}'. "
        f"Based on the sender's original message, draft a polite, professional reply "
        f"that incorporates these instructions:\n\n"
        f"'{user_instructions}'\n\n"
        "Present the reply clearly in a copyable code block."
    )


if __name__ == "__main__":
    mcp.run()